from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from ww.myutil.get_kpi import *
import datetime
import openpyxl
from ww import models
from issue import models as mm
from django.core.paginator import Paginator
import json
from ww.myutil.REDISUtil import my_redis
from pyecharts.charts import Bar,Grid
from pyecharts import options as opts
from pyecharts.commons.utils import JsCode
import re


def g(x):
    if int(x['min']) >= 10 and int(x['min']) <=20:
        return 10


def test_js(request):
    # data = get_all('2022-07-03 00:00:00', '2022-07-04 13:00:00', 'XINMZ78_02B')
    # titles = ['id','IGV号','TOS收箱任务下发时间','任务下发，单车反馈200','卸船单车到达临停位置','卸船单车离开临停位置','卸船单车到达QCTP-X位','卸船单车到达岸桥','卸船单车开始对位','卸船单车结束对位','卸船单车送箱任务完成','单车到达前PB','单车离开前PB','单车到达锁站','单车装锁完成','单车到达后PB','单车离开后PB','箱号1','箱号2','TOS送箱第一箱任务下发时间','任务下发，单车反馈200','单车第一箱到达第一个TP位置','单车第一段开始对位','单车第一段结束对位','单车第一段送箱任务完成','TOS送箱第二箱任务下发时间','任务下发，单车反馈200','单车到达第二个TP位置','单车第二段开始对位','单车第二段结束对位','单车第二段送箱任务完成','总耗电量','ALL_TIME','船舶ID号','SPEED','收箱里程','送箱里程','Task Type','对位用时']
    # new_list = list()
    # for d in data:
    #     new_dict = dict()
    #     for t in titles:
    #         if t == 'ALL_TIME':
    #             s = int(round(float(d[titles.index(t)])/60, 0))
    #             new_dict['min'] = s
    #     new_list.append(new_dict)
    # u_list = sorted(new_list, key=lambda x: x["min"])
    # l_g = groupby(u_list, key=g)
    # for key, group in l_g:
    #     print(key, list(group))
    # c = (
    #     # 设置主题的样式
    #     Bar(init_opts=opts.InitOpts(width="900px", height="500px"))
    #         .add_xaxis(["10", "20", "30", "40", "50", "60"])
    #         .add_yaxis("装船", [27, 44, 29, 19, 10, 4], stack="stack", color="#a834a8")
    #         .add_yaxis("", [2, 11, 20, 9, 8, 24], stack="stack", color="#a834a8")
    #         # 增加主题和副标题
    #         .set_series_opts(
    #         label_opts=opts.LabelOpts(position="inside", color="white", font_size=18, font_style="normal",
    #                                   font_weight='normal', font_family='Times New Roman'))
    #         .set_global_opts(title_opts=opts.TitleOpts(title="作业时间分布", subtitle=""))
    # )
    x = ['10', '15', '20', '25', '30', '35', '40', '45', '50', '55', '60', '65', '70']
    a = [19, 40, 41, 73, 65, 26, 26, 16, 10, 8, 7, 8]
    b = [6, 12, 12, 22, 19, 8, 8, 5, 3, 2, 2, 2]

    bar = Bar()
    bar.add_xaxis(
        xaxis_data=x)
    bar.add_yaxis(
        series_name="装船",
        y_axis=a,
        stack="stack",
        color="#0066cc",
        label_opts=opts.LabelOpts(
            position=[5, -30],
            color="black",
            font_size=14,
            font_style="normal",
            font_weight='normal',
            font_family='Times New Roman',
            is_show=True,
            formatter=JsCode(
                "function(x){return x.data+'圈'}"
            )
        )
    )
    bar.add_yaxis(
        series_name="装船",
        y_axis=b,
        stack="stack",
        color="#0066cc",
        label_opts=opts.LabelOpts(
            position=[5, -40],
            color="black",
            font_size=14,
            font_style="normal",
            font_weight='normal',
            font_family='Times New Roman',
            is_show=True,
            formatter=JsCode(
                "function(x){return x.data+'\n%'}"
            )
        )
    )

    grid = Grid()
    grid.add(bar, grid_opts=opts.GridOpts(pos_top="0%", pos_bottom="0%", pos_left="0%", pos_right="0%"))
    return HttpResponse(bar.render_embed())


def all_redis(request):
    key = request.GET.get('searchParams')
    red = my_redis()
    rs = red.query_all()
    all_l = list()
    print(key,'123123')
    for r in rs:
        if key is None:
            all_r = dict()
            all_r['keys'] = r
            all_l.append(all_r)
        else:
            if re.search(key, r):
                all_r = dict()
                all_r['keys'] = r
                all_l.append(all_r)
    red.close_redis()
    context = {"code": 0, "msg": "", "count": len(all_l), "data": all_l}
    return JsonResponse(context, safe=False)


def get_redis_key(request):
    key = str(request.GET.get('keyParams'))
    if key is not None:
        print(key)
        red = my_redis()
        rs = red.query_list(key)
        print(type(rs))
        r_list = list()
        cnt = 0
        if isinstance(rs, dict):
            for k,v in rs.items():
                r_dict = dict()
                r_dict['k'] = k
                r_dict['v'] = v
                r_list.append(r_dict)
                cnt = 2
        elif isinstance(rs, list):
            for r in rs:
                r_dict = dict()
                print(r[0])
                r_dict['r'] = rs.index(r)
                r_dict['k'] = r[0]
                r_dict['v'] = r[1]
                r_list.append(r_dict)
                cnt = 3
        else:
            r_dict = dict()
            r_dict['v'] = rs
            r_list.append(r_dict)
        print(r_list)
    context = {"code": 0, "msg": "", "count": cnt, "data": r_list}
    return JsonResponse(context, safe=False)


def art_redis(request):
    red = my_redis()
    art_pool = red.query_list('vehicle_pool')
    art_suspense = red.query_list('vehicle_suspense_state')
    art_online = red.query_list('vehicle_online_status')
    art_work_status = red.query_list('wfm_vehicle_work_status')
    art_charge = red.query_list('vehicle_charge_job')
    red.close_redis()
    art_list = list()
    for i in range(1,77):
        art_dict = dict()
        art_str = ''
        if i < 10:
            art_str = 'A00'+ str(i)
        else:
            art_str = 'A0' + str(i)
        art_dict['art'] = art_str

        if art_str in art_pool:
            art_dict['pool'] = eval(art_pool[art_str])['pool_alias']
        else:
            art_dict['pool'] = ''
        art_dict['online'] = art_online[art_str]
        art_dict['work_status'] = art_work_status[art_str]
        if art_str in art_suspense:
            art_dict['suspense'] = art_suspense[art_str]
        else:
            art_dict['suspense'] = ''
        if art_str in art_charge:
            art_dict['charge'] = 'CHARGE'
        else:
            art_dict['charge'] = ''

        art_list.append(art_dict)
    print(art_list)
    context = {"code": 0, "msg": "", "count": len(art_list), "data": art_list}
    return JsonResponse(context, safe=False)


def qc_redis(request):
    red = my_redis()
    ts = red.query_list('wfm_config_qc_ts')
    change = red.query_list('single_qc_change_switch')
    red.close_redis()
    qc_list = list()
    for i in range(301,313):
        qc_dict = dict()
        qc_dict['qc'] = 'Q'+str(i)
        qc_dict['ts'] = ts['Q'+str(i)]
        qc_dict['change'] = change['Q'+str(i)]
        qc_list.append(qc_dict)

    context = {"code": 0, "msg": "", "count": len(qc_list), "data": qc_list}
    return JsonResponse(context, safe=False)


def get_jobs(request):
    limit = int(request.GET.get('limit'))
    page = int(request.GET.get('page'))
    container = request.GET.get('containerId')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    vv_id = request.GET.get('vv_id')
    search = dict()
    if vv_id:
        search['job_content__contains'] = vv_id
    jobs = models.jobs.objects.filter(**search)
    if start_date and end_date:
        start_date_f = datetime.datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S') - datetime.timedelta(hours=8)
        end_date_f = datetime.datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S') - datetime.timedelta(hours=8)
        jobs = jobs.filter(Q(create_on__gte=start_date_f) & Q(create_on__lte=end_date_f))
    paginator = Paginator(jobs, limit)
    js = paginator.page(page)
    page_number = jobs.count()+1
    job_list = list()
    for job in js:
        job_dict = dict()
        job_dict['id'] = job.id
        job_dict['job_id'] = job.job_id
        job_dict['job_status'] = job.job_status
        job_dict['vehicle_id'] = job.vehicle_id
        job_content_dict = json.loads(job.job_content)
        if "vesselVisitID" in job_content_dict.keys():
            job_dict['vesselVisitId'] = job_content_dict['vesselVisitID']
        job_dict['jobType'] = job_content_dict['jobType']
        contains = job_content_dict['plannedContainerDestinationList']
        n = 1
        for c in contains:
            job_dict['con_id'+str(n)] = c['containerNum']
            n = n+1
        job_dict['create_on'] = datetime.datetime.strftime(job.create_on+datetime.timedelta(hours=8), '%Y-%m-%d %H:%M:%S')
        job_dict['update_on'] = datetime.datetime.strftime(job.update_on+datetime.timedelta(hours=8), '%Y-%m-%d %H:%M:%S')
        job_list.append(job_dict)
    context = {"code": 0, "msg": "", "count": page_number, "data": job_list}
    return JsonResponse(context, safe=False)


def get_qc_tl_interval(request):
    limit = int(request.GET.get('limit'))
    page = int(request.GET.get('page'))
    qc_id = request.GET.get('qcid')
    vv_id = request.GET.get('vvid')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    search = dict()
    if qc_id:
        search['qcid'] = qc_id
    if vv_id:
        search['vvid'] = vv_id
    ts = models.qc_tl_interval.objects.filter(**search)
    if start_date and end_date:
        start_date_f = datetime.datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S') - datetime.timedelta(hours=8)
        end_date_f = datetime.datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S') - datetime.timedelta(hours=8)
        ts = ts.filter(Q(update_on__gte=start_date_f) & Q(update_on__lte=end_date_f))
        print(start_date_f,end_date_f)
    paginator = Paginator(ts, limit)
    page_data = paginator.page(page)
    page_number = ts.count()+1
    ts_list = list()
    for t in page_data:
        ts_dict = dict()
        ts_dict['id'] = t.id
        ts_dict['qcid'] = t.qcid
        ts_dict['vvid'] = t.vvid
        ts_dict['start'] = t.start + datetime.timedelta(hours=8)
        ts_dict['end'] = t.end + datetime.timedelta(hours=8)
        ts_dict['interval'] = t.interval / 60
        ts_dict['type'] = t.type
        ts_dict['classification'] = t.classification
        ts_dict['reason'] = t.reason
        ts_dict['description'] = t.description
        ts_list.append(ts_dict)
    context = {"code": 0, "msg": "", "count": page_number, "data": ts_list}
    return JsonResponse(context, safe=False)


def get_abort_data(request):
    limit = int(request.GET.get('limit'))
    page = int(request.GET.get('page'))
    search_dict = dict()

    account = request.GET.get('account')
    vvid = request.GET.get('vvid')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    reason = request.GET.get('reason')

    if account:
        search_dict['account'] = account
    if vvid:
        search_dict['vvid'] = vvid
    if reason:
        search_dict['reason_code'] = reason
    abort_record = models.abort_record.objects.filter(**search_dict)
    if start_date and end_date:
        abort_record = abort_record.filter(Q(create_time__gte=start_date) & Q(create_time__lte=end_date))
    page_number = abort_record.count()+1
    paginator = Paginator(abort_record, limit)
    page_datas = paginator.page(page)

    abort_classification = models.abort_classification.objects.all()
    data = list()
    for ar in page_datas:
        ars = dict()
        for ac in abort_classification:
            if ar.reason_code == ac.code:
                ars['id'] = ar.id
                ars['reason_code'] = ar.reason_code
                ars['vvid'] = ar.vvid
                ars['account'] = ar.account
                ars['vin'] = ar.vin
                ars['reason'] = ac.reason
                ars['belong'] = ac.belong
        data.append(ars)
    context = {"code": 0, "msg": "", "count": page_number, "data": data}
    return JsonResponse(context, safe=False)


def download_kpi(request):
    start_time = request.POST.get('start_date')
    end_time = request.POST.get('end_date')
    v_vid = request.POST.get('v_vid')
    f_name = "KPI%s-%s(%s)" % (start_time[5: 10], end_time[5: 10], v_vid)
    response = HttpResponse(content_type='application/msexcel')
    response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % f_name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DSCH'
    titles = ['id','IGV号','TOS收箱任务下发时间','任务下发，单车反馈200','卸船单车到达临停位置','卸船单车离开临停位置','卸船单车到达QCTP-X位','卸船单车到达岸桥','卸船单车开始对位','卸船单车结束对位','卸船单车送箱任务完成','单车到达前PB','单车离开前PB','单车到达锁站','单车装锁完成','单车到达后PB','单车离开后PB','箱号1','箱号2','TOS送箱第一箱任务下发时间','任务下发，单车反馈200','单车第一箱到达第一个TP位置','单车第一段开始对位','单车第一段结束对位','单车第一段送箱任务完成','TOS送箱第二箱任务下发时间','任务下发，单车反馈200','单车到达第二个TP位置','单车第二段开始对位','单车第二段结束对位','单车第二段送箱任务完成','总耗电量','ALL_TIME','船舶ID号','SPEED','收箱里程','送箱里程','Task Type','对位用时']
    for t in titles:
        ws.cell(1,titles.index(t)+1).value = t

    data = get_all(start_time, end_time, v_vid)
    line_numb = 2
    for line in data:
        cell_numb = 1
        for d in line:
            ws.cell(line_numb, cell_numb).value = d
            cell_numb += 1
        line_numb += 1

    wb.save(response)
    return response


def test(request):
    return render(request, "test.html")


def index(request):
    # return HttpResponse("HELLO WORLD")

    return render(request, "index.html")


def add_record_htm(request):
    return render(request, "page/table/addContent.html")


def my_record(request):
    datas = mm.lcj_record.objects.all()

    return render(request, "recordList.html", {"datas": datas})


def search_record(request):
    title = request.POST.get("title")
    datas = mm.lcj_record.objects.filter(title__contains=title)
    return render(request, "recordList.html", {"datas": datas})


def record_detail_htm(request):
    rc_id = request.GET.get("id")
    datas = mm.lcj_record.objects.filter(id=rc_id)
    return render(request, "page/table/recordDetail.html", {"datas": datas})


def record_update(request):
    rtn_msg = {'code': 20000, "msg": "success"}
    rec_id = int(request.POST.get("id"))
    title = str(request.POST.get("title"))
    content = str(request.POST.get("content"))

    update_time = str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    mm.lcj_record.objects.filter(id=rec_id).update(title=title, content=content, update_time=update_time )
    return JsonResponse(rtn_msg)


def record_update_htm(request):
    rc_id = request.GET.get("id")
    datas = mm.lcj_record.objects.get(id=rc_id)
    data_dict = dict()
    data_dict['id'] = datas.id
    data_dict['title'] = datas.title
    data_dict['content'] = datas.content.replace("\"","\'")
    return render(request, "page/table/updateContent.html", {"datas": data_dict})


def add_record(request):
    rtn_msg = {'code': 10000, "msg": "success"}
    username = str(request.POST.get("username"))
    title = str(request.POST.get("title"))
    content = str(request.POST.get("content"))
    create_time = str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    update_time = str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    mm.lcj_record.objects.create(username=username, title=title, content=content, create_time=create_time, update_time=update_time )
    return JsonResponse(rtn_msg)


def kpi(request):
    start_time = request.GET.get('start_date')
    end_time = request.GET.get('end_date')
    title = ["ID", "VEHICLE_PLATE","TOS_RECEIVE_DISPATCH", "MISSION_RECEIVE_DISPATCH", "VPB_ARRIVE", "VPB_FINISHED","QCTP_X_ARRIVED","QCTP_ARRIVED","QC_ALIGNED_START","QC_ALIGNED_FINISHED","RECEIVE_FINISHED","QPB_ARRIVE","QPB_FINISHED","LOCK_ARRIVED","LOCK_FINISHED","HPB_ARRIVE","HPB_FINISHED","CONTAINER_ID_1","CONTAINER_ID_2","TOS_DELIVER_DISPATCH_1","MISSION_DELIVER_DISPATCH_1","TP_ARRIVED_1","TP_ALIGNED_1_START","TP_ALIGNED_1_FINISHED","TP_FINISHED_1","TOS_DELIVER_DISPATCH_2","MISSION_DELIVER_DISPATCH_2","TP_ARRIVED_2","TP_ALIGNED_2_START","TP_ALIGNED_2_FINISHED","TP_FINISHED_2","BATTERY","ALL_TIME","VISIT_ID","SPEED","RECEIVE_MILEAGE","DELIVER_MILEAGE","TASK_TYPE","QCFT_TIME"]
    data_list = []
    data = get_all(start_time, end_time, '')
    for da in data:
        line = dict()
        for d in da:
            line[title[da.index(d)]] = d
        print(line)
        data_list.append(line)
    ss = {"code": 0, "msg": "", "count": len(data_list), "data": data_list}
    return JsonResponse(ss, safe=False)


def login(request):
    return render(request, "login.html")

