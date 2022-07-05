# @Date    : 2022/6/26
# @Author  : Jin chunlong (chunlong.jin@westwell-lab.com)
# @Project : 
# @File : 
# @Software:

import openpyxl

titles = ['id','IGV号','TOS收箱任务下发时间','任务下发，单车反馈200','卸船单车到达临停位置','卸船单车离开临停位置','卸船单车到达QCTP-X位','卸船单车到达岸桥','卸船单车开始对位','卸船单车结束对位','卸船单车送箱任务完成','单车到达前PB','单车离开前PB','单车到达锁站','单车装锁完成','单车到达后PB','单车离开后PB','箱号1','箱号2','TOS送箱第一箱任务下发时间','任务下发，单车反馈200','单车第一箱到达第一个TP位置','单车第一段开始对位','单车第一段结束对位','单车第一段送箱任务完成','TOS送箱第二箱任务下发时间','任务下发，单车反馈200','单车到达第二个TP位置','单车第二段开始对位','单车第二段结束对位','单车第二段送箱任务完成','总耗电量','ALL_TIME','船舶ID号','SPEED','收箱里程','送箱里程','Task Type','对位用时']


def save_excel_load(list,seqs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DSCH'
    num = 1
    for t in titles:
        ws.cell(1,titles.index(t)+1).value = t
    for v in list:
        if v != '' and list.index(v) > 0:
            line_list = v.split(',')
            for n in seqs:
                if n != 'none':
                    ws.cell(list.index(v)+1,1).value = num
                    ws.cell(list.index(v)+1, seqs.index(n)+2).value = line_list[int(n)]
                else:
                    ws.cell(list.index(v)+1, seqs.index(n)+2).value = ''
            num += 1
    wb.save('KPI.xlsx')
    wb.close()
